
# coding: utf-8

# In[24]:


import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook,load_workbook
import os


# In[25]:


def createWorkbook():
    try:
        # Initialize a workbook 
        wb = Workbook()

        # Get the worksheet in the active workbook
        wb.create_sheet('TeamDetails')

        # Acquire a sheet by its name
        ws = wb['TeamDetails']

        #apply some formatting and write template
        Alig = openpyxl.styles.Alignment(horizontal='center', vertical='center', wrap_text=False)
        FontBold = openpyxl.styles.Font(bold=True)

        ws['A1'] = "TeamName"
        ws['B1'] = "PersonName"
        ws['C1'] = "Gender"
        
        extraSheet = wb['Sheet']
        wb.remove(extraSheet)

        # Enumerate the cells in the second row
        for cell in ws["1:1"]:
            cell.font = FontBold
            cell.alignment = Alig
            
        wb.save('TeamOutput/TeamFormationDetails.xlsx')
    
    except Exception as ex:
        print(ex)


# In[26]:


def readDataFromExcel(file_path = "TeamOutput/TeamFormationDetails.xlsx"):
    try:
        teamdata_excel = pd.read_excel(file_path,sheet_name="TeamDetails")
    except:
        teamdata_excel = pd.read_excel(file_path,sheet_name="teamdetails")
    finally:
        return(teamdata_excel)


# In[27]:


def checkExcelFile(file_path = "TeamOutput/TeamFormationDetails.xlsx"):
    try:
        """Function to check excel file is present or not in a specified folder"""

        if os.path.exists(file_path):
            return True
        else:
            createWorkbook()
            return True
    except Exception as ex:
        return False
        print(ex)


# In[28]:


def check_readExcel(file_path = "TeamOutput/TeamFormationDetails.xlsx"):
    try:
        if checkExcelFile(file_path):
            excel_data = readDataFromExcel(file_path)
            return (excel_data)
        else:
            return
    except:
        print("Not able to read excel file")
        return


# In[29]:


def appendDataToExcel(data_dict, file_loc = "TeamOutput/TeamFormationDetails.xlsx"):
    try:
        """Function to append the data"""
        
        #loading the workook
        wb = load_workbook(file_loc)
        
        #fetching the worksheet
        ws = wb['TeamDetails']

        #convert to dataframe
        data_df = pd.DataFrame.from_dict(data_dict)
        
        data_df = data_df[["TeamName","PersonName","Gender"]]

        # Append the rows of the DataFrame to your worksheet
        for r in dataframe_to_rows(data_df, index=False, header=False):
            ws.append(r)
            
        wb.save(file_loc)
    except Exception as ex:
        print(ex)

